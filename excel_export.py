"""
Builds the personal Excel file the customer downloads.

Same split as the Stockholm sheet's UI / Sheet1: an "Inputs" sheet holds the
customer's values, and the "Results" sheet's cells are written as REAL Excel
formulas (not hardcoded numbers) — if the customer changes a measurement in
Excel, the results recalculate automatically.
"""
import re
from openpyxl import Workbook


def _excel_round_wrap(body: str, rule) -> str:
    if rule in (None, "none"):
        return body
    if rule == "even":
        return f"EVEN({body})"
    if rule == "odd":
        return f"ODD({body})"
    if isinstance(rule, dict):
        if rule.get("type") == "mround":
            return f"MROUND({body},{rule.get('multiple', 1)})"
        if rule.get("type") == "round":
            return f"ROUND({body},{rule.get('digits', 0)})"
    return body


def _substitute_identifiers(expr: str, cell_map: dict) -> str:
    out = expr
    for ident in sorted(cell_map, key=len, reverse=True):
        out = re.sub(rf"\b{re.escape(ident)}\b", cell_map[ident], out)
    return out


def build_workbook(pattern, inputs: dict, results: dict, yarn: dict | None):
    wb = Workbook()

    ws_in = wb.active
    ws_in.title = "Inputs"
    ws_in.append(["Measurement / Input", "Value"])
    ws_in.column_dimensions["A"].width = 46
    ws_in.column_dimensions["B"].width = 16

    cell_map = {}
    all_fields = list(pattern.swatch_fields) + list(pattern.yarn_fields) + list(
        pattern.measurement_fields
    )
    for f in all_fields:
        ws_in.append([f["label"], inputs.get(f["id"], 0)])
        cell_map[f["id"]] = f"Inputs!B{ws_in.max_row}"

    ws_out = wb.create_sheet("Results")
    ws_out.append(["Section", "Stitch / Row Count"])
    ws_out.column_dimensions["A"].width = 40
    ws_out.column_dimensions["B"].width = 16

    for f in pattern.computed_fields:
        ws_out.append([f["label"], None])
        row = ws_out.max_row
        body = _substitute_identifiers(f["formula"], cell_map)
        wrapped = _excel_round_wrap(body, f.get("round"))
        ws_out.cell(row=row, column=2).value = f"={wrapped}"
        cell_map[f["id"]] = f"Results!B{row}"

    if pattern.yarn_estimate and yarn:
        ws_out.append(["Estimated total yarn (meters)", round(yarn["total_meters"], 1)])
        ws_out.append(["Estimated total yarn (grams)", round(yarn["total_weight"], 1)])

    return wb
