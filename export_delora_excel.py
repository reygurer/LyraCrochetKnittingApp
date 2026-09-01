"""
Builds a Delora Top workbook laid out as real Excel Tables in separate
blocks (2x2 Rib swatch, Stockinette swatch, Measurements, Results) on one
sheet — not a plain sequential list. Each block is an actual openpyxl
Table (banded rows, header row, filter buttons), so it reads and behaves
like a spreadsheet table, not just bordered cells.

Delora-specific layout script, kept separate from excel_export.py (which
stays the plain sequential Inputs/Results generator the live app's
"Download as Excel" button uses for every pattern, Aurelia included).
Reuses that module's formula-translation helpers so formulas are produced
the exact same way in both places.

Run: python3 export_delora_excel.py [output_path]
Produces a blank template (Inputs are empty, ready to fill in).
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from delora_pattern import build_delora_pattern
from excel_export import _substitute_identifiers, _excel_round_wrap

TITLE_FONT = Font(bold=True, size=13, color="2B2630")
LETTER_FONT = Font(bold=True, size=10.5, color="2C4D43")
INPUT_FILL = PatternFill("solid", fgColor="FFF9E6")  # pale yellow: "type here"

INPUT_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
)
RESULTS_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium7", showRowStripes=True, showFirstColumn=False
)


def _write_input_table(ws, table_name, top_left, fields, inputs, cell_map):
    """Writes a Field/Value table starting at top_left=(row, col); returns the row after it."""
    row0, col0 = top_left
    label_col, value_col = col0, col0 + 1
    ws.cell(row=row0, column=label_col, value="Field").font = Font(bold=True)
    ws.cell(row=row0, column=value_col, value="Value").font = Font(bold=True)

    row = row0 + 1
    for f in fields:
        ws.cell(row=row, column=label_col, value=f["label"])
        val_cell = ws.cell(row=row, column=value_col, value=inputs.get(f["id"]))
        val_cell.fill = INPUT_FILL
        val_cell.alignment = Alignment(horizontal="center")
        cell_map[f["id"]] = f"{get_column_letter(value_col)}{row}"
        row += 1

    last_row = row - 1
    ref = f"{get_column_letter(label_col)}{row0}:{get_column_letter(value_col)}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = INPUT_TABLE_STYLE
    ws.add_table(tbl)
    return row  # first free row after the table


def build_delora_workbook(inputs: dict | None = None):
    inputs = inputs or {}
    pattern = build_delora_pattern()

    wb = Workbook()
    ws = wb.active
    ws.title = "Delora Top"
    ws.sheet_view.showGridLines = False

    # Column layout: A/B = swatch block (rib, then stockinette below it),
    # D/E = measurements block, G/H/I = results block (letter/description/value).
    for col, width in [(1, 40), (2, 11), (3, 3), (4, 46), (5, 11), (6, 3), (7, 6), (8, 54), (9, 12)]:
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value="DELORA TOP — fill in the yellow cells with your own numbers")
    title_cell.font = TITLE_FONT
    ws.row_dimensions[1].height = 26

    cell_map = {}

    # --- Left column: swatches, stacked -------------------------------------
    row = _write_input_table(ws, "RibSwatch", (3, 1), pattern.swatch_fields[:6], inputs, cell_map)
    _write_input_table(ws, "StockinetteSwatch", (row + 2, 1), pattern.swatch_fields[6:], inputs, cell_map)

    # --- Second column: measurements ----------------------------------------
    _write_input_table(ws, "Measurements", (3, 4), pattern.measurement_fields, inputs, cell_map)

    # --- Right column: results, real Excel formulas, one row per letter -----
    # Sorted A→Z for display, NOT the pattern's own dependency order (see
    # delora_pattern.py's docstring). Two passes make that safe: pass 1
    # reserves every letter's final cell address before any formula text is
    # generated, so a formula can freely reference a letter written further
    # down the sheet — Excel resolves that from its own dependency graph at
    # calc time, it doesn't care what row a formula happens to sit on.
    r_row0 = 3
    ws.cell(row=r_row0, column=7, value="Letter").font = Font(bold=True)
    ws.cell(row=r_row0, column=8, value="Description").font = Font(bold=True)
    ws.cell(row=r_row0, column=9, value="Value").font = Font(bold=True)

    sorted_fields = sorted(pattern.computed_fields, key=lambda f: f["id"])

    for i, f in enumerate(sorted_fields):  # pass 1: reserve addresses
        cell_map[f["id"]] = f"I{r_row0 + 1 + i}"

    for i, f in enumerate(sorted_fields):  # pass 2: write formulas
        r_row = r_row0 + 1 + i
        letter, _, desc = f["label"].partition(" — ")
        ws.cell(row=r_row, column=7, value=letter).font = LETTER_FONT
        ws.cell(row=r_row, column=8, value=desc or f["label"])
        body = _substitute_identifiers(f["formula"], cell_map)
        wrapped = _excel_round_wrap(body, f.get("round"))
        val_cell = ws.cell(row=r_row, column=9, value=f"={wrapped}")
        val_cell.alignment = Alignment(horizontal="center")

    last_results_row = r_row0 + len(sorted_fields)
    results_tbl = Table(displayName="Results", ref=f"G{r_row0}:I{last_results_row}")
    results_tbl.tableStyleInfo = RESULTS_TABLE_STYLE
    ws.add_table(results_tbl)

    return wb


if __name__ == "__main__":
    out = sys.argv[1] if len(sys.argv) > 1 else "delora-top-blank-template.xlsx"
    build_delora_workbook().save(out)
    print("saved:", out)
